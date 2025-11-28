#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""验证Excel文件中的数据格式"""

import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

excel_file = r'D:\skill\beidou\data\processedMAT\csv_output\UTD_processed_latest\UTD_processed_latest_trackData_gpsl1_text.xlsx'

print(f"验证Excel文件: {excel_file}\n")
print("=" * 70)

# 使用openpyxl读取（保留原始格式）
wb = openpyxl.load_workbook(excel_file, data_only=False)
ws = wb.active

print("前10行数据及其类型:\n")

# 读取标题行
headers = [cell.value for cell in ws[1]]
print(f"列名: {headers}\n")

# 读取前10行数据
for row_idx in range(2, 12):  # 行2到11（跳过标题）
    row_data = []
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        value = cell.value
        value_type = type(value).__name__

        if col_idx <= 5:  # 前5列
            row_data.append(f"{value}")
        else:  # 数据列，显示类型
            row_data.append(f"{value} ({value_type})")

    print(f"行 {row_idx-1}: {', '.join(row_data[:7])}")

print("\n" + "=" * 70)
print("关键数据列的类型检查:")
print("=" * 70)

data_cols = ['I_Prompt', 'Q_Prompt', 'CarrierFreq_Hz', 'PLL_Lock', 'FLL_Lock']
col_indices = {headers[i]: i+1 for i in range(len(headers))}

for col_name in data_cols:
    if col_name in col_indices:
        col_idx = col_indices[col_name]
        cell = ws.cell(row=2, column=col_idx)  # 第一行数据
        value = cell.value
        value_type = type(value).__name__

        print(f"{col_name}:")
        print(f"  值: {value}")
        print(f"  类型: {value_type}")
        print(f"  是文本: {isinstance(value, str)}")
        print()

wb.close()

print("=" * 70)
print("结论:")
print("=" * 70)
if isinstance(ws.cell(row=2, column=col_indices['I_Prompt']).value, str):
    print("[OK] 数值已成功转换为文本格式（科学计数法字符串）")
    print("现在用WPS打开应该不会有显示问题了！")
else:
    print("[WARNING] 数值可能仍然是数字格式")
